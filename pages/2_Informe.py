from components.footer import mostrar_pie
import streamlit as st
import sys
sys.path.append(".")
from assets.estilos import aplicar_estilos
aplicar_estilos()
from core.database import run_query
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

st.title("📋 Informe por Programa")
st.markdown("Detalle completo de cada programa académico.")
st.divider()

programas = run_query("SELECT id, nombre, codigo FROM programas ORDER BY nombre")

if not programas:
    st.warning("No se encontraron programas.")
else:
    nombres = [f"{p['nombre']} ({p['codigo']})" for p in programas]
    seleccion = st.selectbox("Selecciona un programa:", nombres)
    prog = programas[nombres.index(seleccion)]

    materias = run_query("""
        SELECT m.id, m.nombre, m.codigo, m.nivel, m.creditos, m.periodo, m.version,
               d.version as version_doc, d.fecha_subida, d.creado_por
        FROM materias m
        LEFT JOIN documentos d ON d.materia_id = m.id AND d.activo = 1
        WHERE m.programa_id = %s
        ORDER BY m.periodo, m.nombre
    """, (prog['id'],))

    total         = len(materias)
    sin_contenido = sum(1 for m in materias if not m['version_doc'])
    con_contenido = total - sin_contenido

    versiones_disponibles = sorted(set(
        m['version_doc'] for m in materias if m['version_doc']
    ))

    num_cols = 3 + len(versiones_disponibles)
    cols = st.columns(num_cols)
    cols[0].metric("Total materias", total)
    cols[1].metric("Con contenido", con_contenido)
    cols[2].metric("Sin contenido", sin_contenido)
    for i, v in enumerate(versiones_disponibles):
        cantidad = sum(1 for m in materias if m['version_doc'] == v)
        cols[3 + i].metric(f"En versión {v}", cantidad)

    st.divider()

    opciones_filtro = ["Todas", "Sin contenido", "Con contenido"] + \
                      [f"En versión {v}" for v in versiones_disponibles]
    filtro = st.radio("Filtrar por:", opciones_filtro, horizontal=True)

    if filtro == "Sin contenido":
        materias_filtradas = [m for m in materias if not m['version_doc']]
    elif filtro == "Con contenido":
        materias_filtradas = [m for m in materias if m['version_doc']]
    elif filtro.startswith("En versión"):
        version_sel = filtro.replace("En versión ", "")
        materias_filtradas = [m for m in materias if m['version_doc'] == version_sel]
    else:
        materias_filtradas = materias

    rol_usuario   = st.session_state.get('usuario_rol', '')
    es_superadmin = (rol_usuario == 'superadmin')

    for m in materias_filtradas:
        estado = "✅ V" + m['version_doc'] if m['version_doc'] else "❌ Sin contenido"
        with st.container(border=True):
            c1, c2, c3, c4 = st.columns([3, 1, 1, 2])
            c1.markdown(f"**{m['nombre']}**  \nCódigo: `{m['codigo'] or 'N/A'}`")
            c2.markdown(f"Semestre: **{m['periodo']}**")
            c3.markdown(f"Créditos: **{m['creditos']}**")
            c4.markdown(f"Estado actual: **{estado}**")

            if es_superadmin:
                historial = run_query("""
                    SELECT version, fecha_subida, creado_por, activo
                    FROM documentos
                    WHERE materia_id = %s
                    ORDER BY fecha_subida DESC
                """, (m['id'],))
                if historial:
                    st.markdown("---")
                    with st.expander("📜 Ver Historial Completo de Cambios (Auditoría)"):
                        for h in historial:
                            fecha_formateada = h['fecha_subida'].strftime('%Y-%m-%d a las %H:%M:%S') if h['fecha_subida'] else "Fecha desconocida"
                            creador = h['creado_por'] if h['creado_por'] else "Sistema / Previos"
                            if h['activo'] == 1:
                                st.markdown(f"🟢 **Versión Actual (V{h['version']}):** Subido por `{creador}` el {fecha_formateada}")
                            else:
                                st.markdown(f"🔘 Reemplazado (V{h['version']}): Cambiado por `{creador}` el {fecha_formateada}")
                else:
                    st.caption("No hay registros de cambios anteriores para esta materia.")

    st.divider()

    # ─────────────────────────────────────────
    # GENERAR PDF
    # ─────────────────────────────────────────
    def encabezado_tabla_pdf(pdf):
        pdf.set_fill_color(45, 90, 27)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(75, 8, "Materia",  border=1, fill=True)
        pdf.cell(25, 8, "Codigo",   border=1, fill=True)
        pdf.cell(18, 8, "Semestre", border=1, fill=True)
        pdf.cell(18, 8, "Creditos", border=1, fill=True)
        pdf.cell(54, 8, "Estado",   border=1, fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "", 9)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 6, "texto", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 11)
    resumen_versiones = "  |  ".join(
        [f"V{v}: {sum(1 for m in materias if m['version_doc'] == v)}"
         for v in versiones_disponibles]
    )
    pdf.multi_cell(0, 8,
        f"Codigo: {prog['codigo']}  |  Total: {total}  |  "
        f"Con contenido: {con_contenido}  |  Sin contenido: {sin_contenido}  |  {resumen_versiones}")
    pdf.ln(4)
    encabezado_tabla_pdf(pdf)

    for idx, m in enumerate(materias):
        estado_txt = "V" + m['version_doc'] if m['version_doc'] else "Sin contenido"
        nombre     = (m['nombre'] or "Sin nombre")[:60]

        if pdf.get_y() > 260:
            pdf.add_page()
            encabezado_tabla_pdf(pdf)

        if idx % 2 == 0:
            pdf.set_fill_color(240, 245, 234)
        else:
            pdf.set_fill_color(255, 255, 255)

        y_inicio = pdf.get_y()
        x_inicio = pdf.get_x()
        pdf.multi_cell(75, 7, nombre, border=1, fill=True)
        y_fin  = pdf.get_y()
        altura = max(y_fin - y_inicio, 7)

        pdf.set_xy(x_inicio + 75, y_inicio)
        pdf.cell(25, altura, str(m['codigo']   or ""), border=1, fill=True)
        pdf.cell(18, altura, str(m['periodo']  or ""), border=1, fill=True)
        pdf.cell(18, altura, str(m['creditos'] or ""), border=1, fill=True)
        pdf.cell(54, altura, estado_txt,                border=1, fill=True)
        pdf.ln()

    pdf_bytes = pdf.output()

    # ─────────────────────────────────────────
    # GENERAR EXCEL
    # ─────────────────────────────────────────
    verde       = "2d5a1b"
    verde_claro = "f0f5ea"
    blanco      = "FFFFFF"
    fill_verde  = PatternFill("solid", fgColor=verde)
    fill_claro  = PatternFill("solid", fgColor=verde_claro)
    borde = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin")
    )
    centrado  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    izquierda = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Informe"

    ws.merge_cells("A1:E1")
    ws["A1"] = f"Informe: {prog['nombre']}"
    ws["A1"].font      = Font(bold=True, size=14, color=blanco)
    ws["A1"].fill      = fill_verde
    ws["A1"].alignment = centrado

    ws.merge_cells("A2:E2")
    resumen = (f"Código: {prog['codigo']}  |  Total: {total}  |  "
               f"Con contenido: {con_contenido}  |  Sin contenido: {sin_contenido}  |  " +
               "  |  ".join([f"V{v}: {sum(1 for m in materias if m['version_doc'] == v)}"
                              for v in versiones_disponibles]))
    ws["A2"] = resumen
    ws["A2"].font      = Font(bold=True, size=10, color=verde)
    ws["A2"].alignment = izquierda

    ws.append([])

    encabezados = ["Materia", "Código", "Semestre", "Créditos", "Estado"]
    ws.append(encabezados)
    for col in range(1, 6):
        cell           = ws.cell(row=4, column=col)
        cell.font      = Font(bold=True, color=blanco)
        cell.fill      = fill_verde
        cell.alignment = centrado
        cell.border    = borde

    for idx, m in enumerate(materias):
        estado_txt = "V" + m['version_doc'] if m['version_doc'] else "Sin contenido"
        ws.append([
            m['nombre']   or "",
            m['codigo']   or "",
            m['periodo']  or "",
            m['creditos'] or "",
            estado_txt
        ])
        fila      = 5 + idx
        fill_fila = fill_claro if idx % 2 == 0 else PatternFill("solid", fgColor=blanco)
        for col in range(1, 6):
            cell           = ws.cell(row=fila, column=col)
            cell.fill      = fill_fila
            cell.border    = borde
            cell.alignment = izquierda if col == 1 else centrado

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # ─────────────────────────────────────────
    # BOTONES DE DESCARGA DIRECTA
    # ─────────────────────────────────────────
    col_pdf, col_excel = st.columns(2)
    with col_pdf:
        st.download_button(
            "📄 Descargar informe PDF",
            data=bytes(pdf_bytes),
            file_name=f"informe_{prog['codigo']}.pdf",
            mime="application/pdf",
            use_container_width=True
        )
    with col_excel:
        st.download_button(
            "📊 Descargar informe Excel",
            data=buffer,
            file_name=f"informe_{prog['codigo']}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

mostrar_pie()